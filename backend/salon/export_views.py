from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone
from django.db.models import Sum, Count
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.views.decorators.http import require_http_methods
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
import io
import json
from datetime import datetime, timedelta
# PDF imports removed - keeping only Excel functionality
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from .models import (
    Booking, Customer, Service, Staff, Category, Coupon, 
    Testimonial, Offer, ContactInfo
)


class ExportMixin:
    """Mixin class for common export functionality"""
    
    def get_date_range(self, request):
        """Get date range from request parameters"""
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        
        if start_date:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        else:
            start_date = timezone.now().date() - timedelta(days=30)
            
        if end_date:
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        else:
            end_date = timezone.now().date()
            
        return start_date, end_date
    
    def get_queryset_filters(self, request):
        """Get common queryset filters"""
        filters = {}
        
        # Date range filter
        start_date, end_date = self.get_date_range(request)
        filters['created_at__date__range'] = [start_date, end_date]
        
        # Status filter
        status_filter = request.GET.get('status')
        if status_filter:
            filters['status'] = status_filter
            
        # Staff filter
        staff_id = request.GET.get('staff_id')
        if staff_id:
            filters['staff_id'] = staff_id
            
        # Service filter
        service_id = request.GET.get('service_id')
        if service_id:
            filters['service_id'] = service_id
            
        return filters


# PDF export function removed - keeping only Excel functionality


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_bookings_excel(request):
    """Export bookings to Excel"""
    try:
        # Get filters
        export_mixin = ExportMixin()
        filters = export_mixin.get_queryset_filters(request)
        
        # Get bookings
        bookings = Booking.objects.filter(**filters).select_related(
            'customer', 'service', 'staff'
        ).order_by('-created_at')
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "تقرير الحجوزات"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Title
        ws.merge_cells('A1:H1')
        ws['A1'] = f"تقرير الحجوزات - صالون الجمال - {timezone.now().strftime('%Y-%m-%d %H:%M')}"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal="center")
        
        # Summary statistics
        total_bookings = bookings.count()
        total_revenue = bookings.aggregate(total=Sum('final_price'))['total'] or 0
        pending_bookings = bookings.filter(status='pending').count()
        completed_bookings = bookings.filter(status='completed').count()
        
        ws['A3'] = "إجمالي الحجوزات:"
        ws['B3'] = total_bookings
        ws['A4'] = "إجمالي الإيرادات:"
        ws['B4'] = f"{total_revenue:.2f} ريال"
        ws['A5'] = "الحجوزات المعلقة:"
        ws['B5'] = pending_bookings
        ws['A6'] = "الحجوزات المكتملة:"
        ws['B6'] = completed_bookings
        
        # Headers
        headers = ['رقم الحجز', 'اسم العميل', 'البريد الإلكتروني', 'رقم الهاتف', 'الخدمة', 'الفئة', 'الموظف', 'التاريخ', 'الوقت', 'الحالة', 'السعر الأصلي', 'مبلغ الخصم', 'السعر النهائي', 'طريقة الدفع', 'تاريخ الإنشاء']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=8, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Data rows
        for row, booking in enumerate(bookings, 9):
            ws.cell(row=row, column=1, value=booking.id)
            ws.cell(row=row, column=2, value=booking.customer.name)
            ws.cell(row=row, column=3, value=booking.customer.email)
            ws.cell(row=row, column=4, value=booking.customer.phone)
            ws.cell(row=row, column=5, value=booking.service.name)
            ws.cell(row=row, column=6, value=booking.service.category.name)
            ws.cell(row=row, column=7, value=booking.staff.name if booking.staff else 'غير محدد')
            ws.cell(row=row, column=8, value=booking.booking_date.strftime('%Y-%m-%d'))
            ws.cell(row=row, column=9, value=booking.booking_time.strftime('%H:%M'))
            ws.cell(row=row, column=10, value=booking.get_status_display())
            ws.cell(row=row, column=11, value=float(booking.price))
            ws.cell(row=row, column=12, value=float(booking.discount_amount))
            ws.cell(row=row, column=13, value=float(booking.final_price))
            ws.cell(row=row, column=14, value=booking.get_payment_method_display())
            ws.cell(row=row, column=15, value=booking.created_at.strftime('%Y-%m-%d %H:%M'))
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="bookings_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# PDF export function removed - keeping only Excel functionality


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_customers_excel(request):
    """Export customers to Excel"""
    try:
        # Get customers
        customers = Customer.objects.all().order_by('-created_at')
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "تقرير العملاء"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Title
        ws.merge_cells('A1:G1')
        ws['A1'] = f"تقرير العملاء - صالون الجمال - {timezone.now().strftime('%Y-%m-%d %H:%M')}"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal="center")
        
        # Summary statistics
        total_customers = customers.count()
        active_customers = customers.filter(is_active=True).count()
        
        ws['A3'] = "إجمالي العملاء:"
        ws['B3'] = total_customers
        ws['A4'] = "العملاء النشطون:"
        ws['B4'] = active_customers
        ws['A5'] = "العملاء غير النشطين:"
        ws['B5'] = total_customers - active_customers
        
        # Headers
        headers = ['رقم العميل', 'الاسم', 'البريد الإلكتروني', 'رقم الهاتف', 'تاريخ الميلاد', 'الحالة', 'تاريخ التسجيل']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=7, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Data rows
        for row, customer in enumerate(customers, 8):
            ws.cell(row=row, column=1, value=customer.id)
            ws.cell(row=row, column=2, value=customer.name)
            ws.cell(row=row, column=3, value=customer.email)
            ws.cell(row=row, column=4, value=customer.phone)
            ws.cell(row=row, column=5, value=customer.date_of_birth.strftime('%Y-%m-%d') if customer.date_of_birth else 'غير محدد')
            ws.cell(row=row, column=6, value='نشط' if customer.is_active else 'غير نشط')
            ws.cell(row=row, column=7, value=customer.created_at.strftime('%Y-%m-%d'))
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="customers_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# PDF export function removed - keeping only Excel functionality


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_services_excel(request):
    """Export services to Excel"""
    try:
        # Get services
        services = Service.objects.filter(is_active=True).order_by('name')
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "تقرير الخدمات"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Title
        ws.merge_cells('A1:G1')
        ws['A1'] = f"تقرير الخدمات - صالون الجمال - {timezone.now().strftime('%Y-%m-%d %H:%M')}"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal="center")
        
        # Summary statistics
        total_services = services.count()
        featured_services = services.filter(is_featured=True).count()
        
        ws['A3'] = "إجمالي الخدمات:"
        ws['B3'] = total_services
        ws['A4'] = "الخدمات المميزة:"
        ws['B4'] = featured_services
        ws['A5'] = "الخدمات العادية:"
        ws['B5'] = total_services - featured_services
        
        # Headers
        headers = ['رقم الخدمة', 'اسم الخدمة', 'اسم الخدمة (إنجليزي)', 'الفئة', 'السعر', 'المدة', 'مميزة', 'تاريخ الإنشاء']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=7, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Data rows
        for row, service in enumerate(services, 8):
            ws.cell(row=row, column=1, value=service.id)
            ws.cell(row=row, column=2, value=service.name)
            ws.cell(row=row, column=3, value=service.name_en)
            ws.cell(row=row, column=4, value=service.category.name)
            ws.cell(row=row, column=5, value=float(service.price))
            ws.cell(row=row, column=6, value=service.duration)
            ws.cell(row=row, column=7, value='نعم' if service.is_featured else 'لا')
            ws.cell(row=row, column=8, value=service.created_at.strftime('%Y-%m-%d'))
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="services_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# PDF export function removed - keeping only Excel functionality


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_revenue_report_excel(request):
    """Export revenue report to Excel"""
    try:
        # Get date range
        export_mixin = ExportMixin()
        start_date, end_date = export_mixin.get_date_range(request)
        
        # Get bookings in date range
        bookings = Booking.objects.filter(
            booking_date__range=[start_date, end_date]
        ).select_related('service')
        
        # Calculate revenue statistics
        total_revenue = bookings.aggregate(total=Sum('final_price'))['total'] or 0
        total_discounts = bookings.aggregate(total=Sum('discount_amount'))['total'] or 0
        total_bookings = bookings.count()
        avg_booking_value = total_revenue / total_bookings if total_bookings > 0 else 0
        
        # Revenue by service category
        revenue_by_category = {}
        for booking in bookings:
            category_name = booking.service.category.name
            if category_name not in revenue_by_category:
                revenue_by_category[category_name] = 0
            revenue_by_category[category_name] += float(booking.final_price)
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "تقرير الإيرادات"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Title
        ws.merge_cells('A1:B1')
        ws['A1'] = f"تقرير الإيرادات - صالون الجمال - {start_date} إلى {end_date}"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal="center")
        
        # Summary statistics
        ws['A3'] = "إجمالي الإيرادات:"
        ws['B3'] = f"{total_revenue:.2f} ريال"
        ws['A4'] = "إجمالي الخصومات:"
        ws['B4'] = f"{total_discounts:.2f} ريال"
        ws['A5'] = "عدد الحجوزات:"
        ws['B5'] = total_bookings
        ws['A6'] = "متوسط قيمة الحجز:"
        ws['B6'] = f"{avg_booking_value:.2f} ريال"
        
        # Revenue by category
        if revenue_by_category:
            ws['A8'] = "الإيرادات حسب الفئة:"
            ws['A8'].font = Font(bold=True, size=14)
            
            # Headers
            ws['A9'] = "الفئة"
            ws['B9'] = "الإيرادات"
            
            for col in [9]:
                cell = ws.cell(row=col, column=1)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell = ws.cell(row=col, column=2)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Data
            row = 10
            for category, revenue in revenue_by_category.items():
                ws.cell(row=row, column=1, value=category)
                ws.cell(row=row, column=2, value=f"{revenue:.2f} ريال")
                row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="revenue_report_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_dashboard_data(request):
    """Export comprehensive dashboard data"""
    try:
        # Get date range
        export_mixin = ExportMixin()
        start_date, end_date = export_mixin.get_date_range(request)
        
        # Get comprehensive data
        bookings = Booking.objects.filter(
            booking_date__range=[start_date, end_date]
        ).select_related('customer', 'service', 'staff')
        
        customers = Customer.objects.filter(
            created_at__date__range=[start_date, end_date]
        )
        
        # Calculate statistics
        stats = {
            'total_bookings': bookings.count(),
            'total_revenue': bookings.aggregate(total=Sum('final_price'))['total'] or 0,
            'total_customers': customers.count(),
            'avg_booking_value': 0,
            'bookings_by_status': {},
            'revenue_by_category': {},
            'bookings_by_staff': {},
            'daily_revenue': {}
        }
        
        # Calculate average booking value
        if stats['total_bookings'] > 0:
            stats['avg_booking_value'] = stats['total_revenue'] / stats['total_bookings']
        
        # Bookings by status
        for status_choice in Booking.STATUS_CHOICES:
            status_code = status_choice[0]
            count = bookings.filter(status=status_code).count()
            stats['bookings_by_status'][status_choice[1]] = count
        
        # Revenue by category
        for booking in bookings:
            category_name = booking.service.category.name
            if category_name not in stats['revenue_by_category']:
                stats['revenue_by_category'][category_name] = 0
            stats['revenue_by_category'][category_name] += float(booking.final_price)
        
        # Bookings by staff
        for booking in bookings:
            staff_name = booking.staff.name if booking.staff else 'غير محدد'
            if staff_name not in stats['bookings_by_staff']:
                stats['bookings_by_staff'][staff_name] = 0
            stats['bookings_by_staff'][staff_name] += 1
        
        # Daily revenue
        current_date = start_date
        while current_date <= end_date:
            daily_bookings = bookings.filter(booking_date=current_date)
            daily_revenue = daily_bookings.aggregate(total=Sum('final_price'))['total'] or 0
            stats['daily_revenue'][current_date.strftime('%Y-%m-%d')] = {
                'bookings': daily_bookings.count(),
                'revenue': float(daily_revenue)
            }
            current_date += timedelta(days=1)
        
        return JsonResponse({
            'success': True,
            'data': stats,
            'date_range': {
                'start_date': start_date.strftime('%Y-%m-%d'),
                'end_date': end_date.strftime('%Y-%m-%d')
            }
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
